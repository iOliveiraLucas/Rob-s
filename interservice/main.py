""" main.py """

# Libraries
import requests
from time import sleep
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
import urllib
import base64
from datetime import datetime as dt
from datetime import timedelta as td
from datetime import date
from sqlalchemy import create_engine
import pandas as pd
import schedule
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import openpyxl

chrome_options = Options()
# chrome_options.add_argument('--headless') #enabled in productions
chrome_options.add_argument('--no-sandbox') #enabled in productions
chrome_options.add_argument("--disable-setuid-sandbox") #enabled in productions
chrome_options.add_argument("window-size=1200x600") #enabled in productions
chrome = webdriver.Chrome(chrome_options=chrome_options, executable_path="C:/Users/alexs/Desktop/robos/interservice/chromedriver")#'/usr/bin/chromedriver')

data = {
    'Login': '',
    'Senha': '',
    'Captcha': '',
    'url_login':'',
    'url_logged':''
}

def get_ids():
    db_conn = {'user': '', 'password': '', 'url': '', 'db': '', 'port': '', 'driver': ''}
    engine = create_engine(db_conn['driver']+'://'+db_conn['user']+':'+db_conn['password']+'@'+db_conn['url']+':'+db_conn['port']+'/'+db_conn['db'],echo=False)
    dataframe = pd.read_sql("SELECT NumeroConsulta FROM tabela WHERE length(NumeroConsulta) = 7 GROUP BY NumeroConsulta ORDER BY NumeroConsulta ASC", engine)
    data = dataframe.values.tolist()
    return data

def login():
    logged = {}
    logged['status'] = False

    image = chrome.find_elements_by_xpath('//*[@id="form1"]/div[3]/section/div/section[2]/div[2]/fieldset/div[4]/div/span[1]/img')
    captcha = image[0].get_attribute("src")
    urllib.request.urlretrieve(captcha, "local.jpg")
    
    raw = open("local.jpg", 'rb').read()
    data['Captcha'] = solver2captcha(raw_image=raw)
    chrome.find_element_by_name("ctl00$cphConteudo$txtLogin").send_keys(data["Login"])
    chrome.find_element_by_name("ctl00$cphConteudo$txtSenha").send_keys(data["Senha"])
    chrome.find_element_by_name("ctl00$cphConteudo$CaptchaFaleConosco").send_keys(data["Captcha"])
    chrome.find_element_by_id("btnAcessar").click()
    chrome.get(url=data['url_logged'])

    logged['status'] = True
    return logged
    
def search_ids():
    list_ids = get_ids()
    chrome.get(url=data['url_login'])
    log_in = login()

    if log_in['status']:
        valid_time = td(minutes=+30) + dt.now()
    lista = []
    
    for item in list_ids:
        validation = False
        hr_now = dt.now()
        
        if hr_now <= valid_time:
            try:
                lista2 = []
                chrome.find_element_by_name("ctl00$cphConteudo$txtFilterIdPropostaCrim").clear()
                chrome.find_element_by_id("rdoSixMonths").click()
                sleep(0.3)
                chrome.find_element_by_name("ctl00$cphConteudo$txtFilterIdPropostaCrim").send_keys(str(item))
                chrome.find_element_by_name("ctl00$cphConteudo$btnFiltrar").click()
                sleep(1.5)
                try:
                    chrome.find_element_by_id('cphConteudo_listViewProposta_ctrl0_lnkBtnDetails_0').click()
                    sleep(0.6)
                    id = chrome.find_elements_by_css_selector('#lblTitleProposalNumber')[0].text
                    cpf = chrome.find_element_by_xpath('/html/body/div[2]/form/div[4]/div[1]/section/div/section[3]/div/div[1]/table/tbody/tr/td/div/div[2]/div[2]/div[2]').get_attribute('innerText')
                    name = chrome.find_elements_by_css_selector('#lblClientName')[0].text
                    result = chrome.find_element_by_xpath('/html/body/div[2]/form/div[4]/div[1]/section/div/section[3]/div/div[2]/div/h3')
                    
                    lista2.append(id)
                    lista2.append(cpf)
                    lista2.append(name)
                    lista2.append(result) 
                    lista2.append(str(date.today()))

                    lista.append(lista2)

                    sleep(0.5)
                    chrome.back()
                    print(id+" - "+cpf+" - "+name+": "+result+' / '+str(date.today()))
                except:
                    pass
            except:
                chrome.get(url=data['url_logged'])
        else:
            chrome.find_element_by_id("lnkSair").click()
            sleep(1)
            validation = False
            while validation != True:
               response_recaptcha2 = login()
               if response_recaptcha2['status'] == True:
                   valid_time = td(minutes=+30) + dt.now()
                   validation = True

    chrome.find_element_by_id("lnkSair").click()
        
    filter = filtering(lista)
    excel = to_excel(lista, filter)
    if excel['status']:
        send_email = email()
        if send_email['status']:
            print('email_enviado!')
            remove_old()
    
def solver2captcha(raw_image):
    base64_binary = base64.b64encode(raw_image)
    base64_text = base64_binary
  
    if len(str(base64_binary).split("'")) == 3:
        base64_text = str(base64_binary).split("'")[1]

    data = {
        'method': 'base64',
        'key': '',
        'body': base64_text
    }
  
    if data['body']:
        response1 = requests.post('https://2captcha.com/in.php', data=data)
        
    if response1:
      result = {}
      track = {}
      if len(response1.text.split("|")) == 2:
        track['status'] = response1.text.split('|')[0]
        track['id'] = response1.text.split('|')[1]
        params = (
          ('key', ''),
          ('action', 'get'),
          ('id', track['id'])
        )

        count = 0
        while True and count < 10:
            response2 = requests.get('https://2captcha.com/res.php', params=params)
            if response2.text != 'CAPCHA_NOT_READY':
              break
            else:
              count = count + 1
            print('Captcha :',response2.text)
            sleep(1)

        if len(response2.text.split("|")) == 2:
            result['status'] = response2.text.split('|')[0]
            result['result'] = response2.text.split('|')[1]
            if result['status'] == 'OK':
                return result['result']
            else:
                print(response1.text)

def to_excel(complete, filtered):
    excel = {}
    excel_complete = pd.DataFrame(data=complete, columns=['Numero Proposta', 'CPF', 'Cliente', 'Situação', 'Data'])
    excel_filtered = pd.DataFrame(data=filtered, columns=['Numero Proposta', 'CPF', 'Cliente', 'Situação', 'Data'])

    with pd.ExcelWriter('./spreadsheets/spreadsheet_interservice_'+str(date.today())+'.xlsx') as writer:
        excel_complete.to_excel(writer, sheet_name='Complete',index=False)
        excel_filtered.to_excel(writer, sheet_name='Filtered',index=False)

    excel['status'] = True
    return excel

def email():
    email = {}
    login = 'robo@email.com.br'
    password = 'Bot@Mail'
    host = 'smtp.gmail.com'
    port = ''
    corpo = 'Planilha InterService ('+str(date.today())+')'

    server = smtplib.SMTP(host,port)
    server.ehlo()
    server.starttls()
    server.login(login,password)

    email_msg = MIMEMultipart()
    to = ['']
    cc = ['']
    email_msg['From'] = login
    email_msg['To'] = ','.join(to) 
    email_msg['Cc'] = ','.join(cc)
    to_addrs = to + cc
    email_msg['Subject'] = 'Planilha InterService ('+str(date.today())+')'
    email_msg.attach(MIMEText(corpo, 'plain'))

    path_sheet = './spreadsheets/spreadsheet_interservice_'+str(date.today())+'.xlsx'
    attachment = open(path_sheet,'rb')
    att = MIMEBase('application', 'octet-stream')
    att.set_payload(attachment.read())
    encoders.encode_base64(att)

    att.add_header('Content-Disposition', f'attachment; filename= spreadsheet_interservice_'+str(date.today())+'.xlsx')
    attachment.close()
    email_msg.attach(att)

    server.sendmail(email_msg['From'],to_addrs,email_msg.as_string())
    server.quit()

    email['status'] = True
    return email

def remove_old():
    old_date = str(td(days=-2)+date.today())
    file_name = 'spreadsheet_interservice_'+old_date
    try:
        os.remove('./spreadsheets/'+file_name+'.xlsx')
    except:
        pass

def filtering(data):
    file_path = r'spreadsheets/spreadsheet_interservice_'+str(td(days=-1)+date.today())+'.xlsx'
    list_filtered = []
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path).active
        list_excel = []
        list_ids = []
        list_status = []
        
        for i in range(1, wb.max_row+1):
            list_rows = []
            for j in range(1, wb.max_column+1):
                cell_obj = wb.cell(row=i, column=j)
                list_rows.append(cell_obj.value)

            list_excel.append(list_rows), list_ids.append(list_rows[0])
        list_excel.pop(0), list_rows.pop(0)

        for item in data:
            if 'Seu contrato foi registrado!' in item:
                list_status.append(item)
                
        for item in list_excel:
            for item2 in list_status:
                if (item[0] == item2[0] and item[3] != item2[3]):
                    list_filtered.append(item2)
                elif (item2[0] not in list_ids) and (item2 not in list_filtered):
                    list_filtered.append(item2)
    else:
        for item in data:
            if 'Seu contrato foi registrado!' in item:
                list_filtered.append(item)
                
    return list_filtered

schedule.every().day.at("06:00").do(search_ids)

if __name__ == '__main__':
    search_ids()
    while 1:
        schedule.run_pending()
        sleep(1)