''' util.py '''

# Libraries
import openpyxl
from datetime import timedelta as td
from datetime import datetime as dt
from datetime import date
from time import sleep
import os

def remove_special_chars(item):
    temp_replaced= str(item).replace('R$','')
    temp_replaced = temp_replaced.replace('.','')
    replaced = temp_replaced.replace(',','.')
    return replaced

def is_loaded(chrome, By):
    loaded = False
    while loaded == False:
        try:
            loading = chrome.find_element(By.XPATH, '//div[@style="display: none;"]//div[@class="loader"]')
            if loading:
                loaded = True
        except:
            loaded = False
            sleep(5)
    return loaded

def rename_file():
    for file in os.listdir('./'):
        if "Relat" in file:
            source = './' + file
            dest = './Planilha.xlsx'
            os.rename(source,dest)

def get_date(days):
    data = str(date.today()-td(days=days))
    return data

def convert_date(temp_date):
    temp_date = str(temp_date)

    if temp_date != 'N/A':
        temp_var = dt.strptime(temp_date, '%d/%m/%Y').date()
        date_convert = "'"+str(temp_var)+"'"
    else:
        date_convert = 'null'

    return date_convert

def remove_old():
    try:
        os.remove('./Planilha.xlsx')
    except:
        pass

def read_excel():
    list_excel = []
    wb = openpyxl.load_workbook('./Planilha.xlsx').active
    
    for i in range(1, wb.max_row+1):
        list_rows = []
        for j in range(1, wb.max_column+1):
            cell_obj = wb.cell(row=i, column=j)
            list_rows.append(cell_obj.value)
        list_excel.append(list_rows)
    list_excel.pop(0)
    
    return list_excel

def convert_dict(temp_list):
    list_reversed= []
    list_data = []

    for item in range(len(temp_list)):
        temp= temp_list.pop(-1)
        list_reversed.append(temp)

    keys = ('numero_proposta','status','primeiro_comprador','cpf','corban','pdv','valor_financiamento','taxa_anual','data_envio','data_emissao','data_assinatura','proposta_enviada_por','produto')
    for item in list_reversed:
        temp_list_data = {keys[i]:item[i] for i in range(13)}
        list_data.append(temp_list_data)
    return list_data

def normalization(item):
    item = str(item)
    # padrão do excel (...) => padrão do db (...)
    # [(N/A),(23/12/2021)]
    if '/' in item:
        if 'N/A' in item:
            var = None  # (N/A) => (None)
        else:
            var = str(dt.strptime(item, '%d/%m/%Y').date())  # (23/12/2021) => (23-12-2021)

    # [(R$65.000,00),(9.5)]
    elif '.' in item:
        if 'R$' in item:
            var = remove_special_chars(item)  # (R$65.000,00) => (65000.00)
        else:
            num = item.replace(".","")
            if num.isnumeric():
                num_int, num_dec = str(item).split('.')
                while len(num_dec) != 5:
                    num_dec += '0'
                var = num_int + '.' + num_dec  # (9.5) => (9.50000) || (17.65) => (17.65000) //5 casas apos o '.'
            else:
                var = item  # entra no if por ter '.' mas não precisa ser alterado

    # [(05447397),(7)]
    elif item.isnumeric():
        if len(item) == 8:
            var = int(item)
            var = str(var)  # (05447397) => (5447397)
        elif(len(item) == 1 or len(item) == 2):
            var = item+'.00000'  # (7) => (7.00000) || (10) => (10.00000)
        else:
            var = item  # entra no if por numerico mas não precisa ser alterado

    # [não precisa ser alterado]
    else:
        var = item
    return var